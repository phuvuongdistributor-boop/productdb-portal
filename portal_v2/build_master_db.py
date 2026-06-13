from __future__ import annotations

import argparse
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MASTER_DIR = PROJECT_ROOT / "master_v2"
DEFAULT_HOTLINE = "0929878666"
DROP_COLUMNS = {"Template_ID", "TheOne_URL"}


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def source_files() -> list[Path]:
    files = [
        path
        for path in PROJECT_ROOT.glob("*/*_FINAL_V2.xlsx")
        if not path.name.startswith("~$")
        and path.name not in {"MASTER_DB.xlsx", "MASTER_DB_AUDIT.xlsx"}
        and "AUDIT" not in path.parts
        and "BACKUP" not in path.parts
    ]
    return sorted(files)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def normalize_workbook(source: Path, destination: Path) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(source)
    worksheet = workbook[workbook.sheetnames[0]]
    original_headers = [cell.value for cell in worksheet[1]]
    kept_indexes = [
        index for index, header in enumerate(original_headers, start=1)
        if header not in DROP_COLUMNS
    ]
    headers = [str(original_headers[index - 1]) for index in kept_indexes]

    if "Source_Group" not in headers:
        headers.append("Source_Group")
    if "Source_URL" not in headers:
        headers.append("Source_URL")

    output = Workbook()
    target = output.active
    target.title = "MASTER_V2"

    header_map = {str(value): index for index, value in enumerate(original_headers, start=1)}
    for target_col, header in enumerate(headers, start=1):
        target_cell = target.cell(1, target_col, header)
        if header in header_map:
            copy_cell_style(worksheet.cell(1, header_map[header]), target_cell)
        else:
            target_cell.font = Font(bold=True, color="FFFFFF")
            target_cell.fill = PatternFill("solid", fgColor="17365D")
            target_cell.alignment = Alignment(horizontal="center")

    output_rows: list[list[object]] = []
    source_group = source.parent.name
    for source_row in worksheet.iter_rows(min_row=2):
        values = [source_row[index - 1].value for index in kept_indexes]
        if all(is_blank(value) for value in values):
            continue

        row_by_header = dict(zip([str(original_headers[i - 1]) for i in kept_indexes], values))
        original_row = {
            str(header): source_row[index - 1].value
            for index, header in enumerate(original_headers, start=1)
        }
        if is_blank(row_by_header.get("Hotline")):
            row_by_header["Hotline"] = DEFAULT_HOTLINE
        row_by_header["Source_Group"] = source_group
        search_url = original_row.get("Search_URL")
        theone_url = original_row.get("TheOne_URL")
        row_by_header["Source_URL"] = search_url if not is_blank(search_url) else theone_url
        normalized = [row_by_header.get(header) for header in headers]
        output_rows.append(normalized)

        target_row = target.max_row + 1
        for target_col, (header, value) in enumerate(zip(headers, normalized), start=1):
            target_cell = target.cell(target_row, target_col, value)
            if header in header_map:
                copy_cell_style(worksheet.cell(source_row[0].row, header_map[header]), target_cell)

    target.freeze_panes = "A2"
    target.auto_filter.ref = target.dimensions
    for index, header in enumerate(headers, start=1):
        original_index = header_map.get(header)
        width = worksheet.column_dimensions[get_column_letter(original_index)].width if original_index else None
        target.column_dimensions[get_column_letter(index)].width = width or min(max(len(header) + 2, 12), 45)

    destination.parent.mkdir(parents=True, exist_ok=True)
    output.save(destination)
    workbook.close()
    output.close()
    return headers, output_rows


def write_master(headers: list[str], rows: list[list[object]]) -> Path:
    path = MASTER_DIR / "MASTER_DB.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "MASTER_DB"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, header in enumerate(headers, start=1):
        sample = [str(row[index - 1]) for row in rows[:500] if not is_blank(row[index - 1])]
        width = min(max([len(header), *[len(value) for value in sample]] if sample else [len(header)]) + 2, 55)
        worksheet.column_dimensions[get_column_letter(index)].width = max(width, 12)

    workbook.save(path)
    workbook.close()
    return path


def write_audit(headers: list[str], rows: list[list[object]], defaulted_hotlines: int) -> Path:
    index = {header: position for position, header in enumerate(headers)}

    def missing(column: str) -> int:
        return sum(1 for row in rows if is_blank(row[index[column]]))

    codes = {str(row[index["Code"]]).strip() for row in rows if not is_blank(row[index["Code"]])}
    groups = Counter(
        str(row[index["Source_Group"]]).strip() or "(Blank)" for row in rows
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    metrics = [
        ("Total Rows", len(rows)),
        ("Unique Code", len(codes)),
        ("Missing Image_URL", missing("Image_URL")),
        ("Missing SalePrice", missing("SalePrice")),
        ("Missing Hotline", missing("Hotline")),
        ("Hotline Filled With Default", defaulted_hotlines),
        (
            "Unresolved Local Image Path",
            sum(
                1
                for row in rows
                if not is_blank(row[index["Image_URL"]])
                and not str(row[index["Image_URL"]]).strip().lower().startswith(("http://", "https://"))
                and not (PROJECT_ROOT / str(row[index["Image_URL"]])).is_file()
            ),
        ),
    ]
    for metric in metrics:
        summary.append(metric)

    by_group = workbook.create_sheet("Rows_by_Source_Group")
    by_group.append(["Source_Group", "Rows"])
    for group, count in sorted(groups.items()):
        by_group.append([group, count])

    for worksheet in (summary, by_group):
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17365D")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 32
        worksheet.column_dimensions["B"].width = 18

    path = MASTER_DIR / "MASTER_DB_AUDIT.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def build() -> dict[str, object]:
    files = source_files()
    if not files:
        raise FileNotFoundError("No *_FINAL_V2.xlsx source files were found.")

    MASTER_DIR.mkdir(parents=True, exist_ok=True)
    all_headers: list[str] | None = None
    all_rows: list[list[object]] = []
    created: list[str] = []
    defaulted_hotlines = 0

    for source in files:
        destination = MASTER_DIR / source.name.replace("_FINAL_V2.xlsx", "_MASTER_V2.xlsx")
        headers, rows = normalize_workbook(source, destination)
        if all_headers is None:
            all_headers = headers
        elif headers != all_headers:
            raise ValueError(f"Column mismatch in {source}")
        hotline_index = headers.index("Hotline")
        defaulted_hotlines += sum(1 for row in rows if str(row[hotline_index]) == DEFAULT_HOTLINE)
        all_rows.extend(rows)
        created.append(destination.name)

    assert all_headers is not None
    master = write_master(all_headers, all_rows)
    audit = write_audit(all_headers, all_rows, defaulted_hotlines)
    return {
        "source_files": len(files),
        "total_rows": len(all_rows),
        "master": str(master),
        "audit": str(audit),
        "created_master_files": created,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build ProductDB V2 master workbooks.")
    parser.parse_args()
    result = build()
    print(f"Built {result['source_files']} source files and {result['total_rows']} SKU rows.")
    print(f"Master: {result['master']}")
    print(f"Audit: {result['audit']}")


if __name__ == "__main__":
    main()
