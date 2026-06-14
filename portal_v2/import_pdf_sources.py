from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from urllib.parse import quote_plus

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PDF_DIR = Path(r"C:\Users\Admin\Desktop\DO")
IMAGE_ROOT = PROJECT_ROOT / "assets" / "product_images"
HOTLINE = "0929878666"
HEADERS = [
    "Code", "ProductName", "Size", "Category", "Variant", "BasePrice",
    "SaleFactor", "SalePrice", "Status", "Search_URL", "Search_Keyword",
    "Material", "SubCategory", "Description", "CatalogPrice", "Price_Mode",
    "Hotline", "Image_Status", "Image_URL", "Image_Source", "Harvest_Status",
    "Source_Group", "Source_URL",
]


T8_PRODUCTS = [
    ("4001", "Đen", 255_000), ("B-140", "Nâu", 1_200_000),
    ("883", "Đen", 255_000), ("C-88", "Đen", 1_250_000),
    ("4019", "Đen", 260_000), ("698A", "Đen", 1_500_000),
    ("4009C", "Đen", 275_000), ("J098A", "Đen", 1_700_000),
    ("4020", "Đen", 270_000), ("A38", "Đen", 1_800_000),
    ("4010C", "Đen", 290_000), ("FY61-1", "Đen/Nâu", 1_900_000),
    ("8008", "Đen", 340_000), ("4005", "Đen", 340_000),
    ("333A", "Nâu", 2_150_000),
    ("1001", "Đen", 390_000), ("TXW-5039", "Đen/Nâu", 2_100_000),
    ("6019", "Sọc đen trắng", 420_000), ("1003", "Đen", 400_000),
    ("A39", "Đen", 2_250_000), ("6020B", "Đen da", 420_000),
    ("1666A", "Nâu/Xám", 2_250_000), ("5003", "Đen", 440_000),
    ("5004", "Đen", 460_000), ("001", "Đen", 2_600_000),
    ("981P", "Đen", 580_000), ("FY90-5", "Đen", 2_600_000),
    ("2006A", "Đen/Vàng", 700_000), ("5037", "Trắng cam/Đen gỗ", 2_600_000),
    ("W118", "Nâu/Đen", 700_000), ("618", "Xám", 2_350_000),
    ("9012", "Xám", 1_020_000), ("5035", "Ghi gỗ", 2_450_000),
    ("W090", "Đen", 1_000_000), ("5038", "Xám cam", 2_700_000),
    ("W070", "Đen", 1_100_000), ("A203", "Xám ghi", 3_300_000),
    ("A335", "Xanh mint", 2_600_000), ("245", "Đen", 2_550_000),
    ("366", "Đen", 1_500_000), ("A222", "Nâu", 4_800_000),
    ("2001", "Đen", 550_000), ("HT-533", "Đen", 450_000),
    ("2002", "Đen", 570_000), ("2005", "Đen", 710_000),
    ("2006", "Đen", 720_000), ("2003", "Đen", 580_000),
    ("2007", "Đen", 720_000), ("2004", "Đen", 600_000),
    ("2008", "Đen", 740_000),
]

T8_IMAGE_XREFS = {
    "4001": [20], "B-140": [28], "883": [22], "C-88": [29],
    "4019": [21], "698A": [34], "4009C": [23], "J098A": [30],
    "4020": [24], "A38": [31], "4010C": [25], "FY61-1": [35, 32],
    "8008": [26], "4005": [27], "333A": [33], "1001": [40],
    "TXW-5039": [46, 47], "6019": [39], "1003": [41], "A39": [48],
    "6020B": [38], "1666A": [49, 55], "5003": [42], "5004": [43],
    "001": [50], "981P": [44], "FY90-5": [51], "2006A": [45, 54],
    "5037": [52, 53], "W118": [79, 80], "618": [66], "9012": [63],
    "5035": [67], "W090": [64], "5038": [68], "W070": [65],
    "A203": [69], "A335": [71], "245": [72], "366": [77],
    "A222": [70], "2001": [76], "HT-533": [78], "2002": [74],
    "2005": [73], "2006": [75], "2003": [83], "2007": [85],
    "2004": [84], "2008": [86],
}


GEVI_PRODUCTS = [
    ("GX4005", "Ghế xoay lưới 4005", "56x56x86-98 cm", 360_000),
    ("GX1001", "Ghế xoay lưới lưng chữ A", "49x50x88-98 cm", 400_000),
    ("GX1003", "Ghế xoay lưng thấp", "60x47x88-98 cm", 490_000),
    ("GX5003", "Ghế xoay tựa đầu", "60x51x107-117 cm", 500_000),
    ("GX2005", "Ghế xoay lưới lưng thấp", "58x50x88-93 cm", 620_000),
    ("GX2006", "Ghế xoay lưới lưng cao", "58x52x108-115 cm", 650_000),
    ("GX6011", "Ghế xoay tay vuông lưng thấp", "58x50x88-93 cm", 630_000),
    ("GX6012", "Ghế xoay tay vuông lưng cao", "58x52x108-115 cm", 660_000),
    ("GX2007", "Ghế xoay da lưng thấp", "58x50x88-93 cm", 690_000),
    ("GX2008", "Ghế xoay da lưng cao", "58x52x108-115 cm", 720_000),
    ("GX528/D1-528AB", "Ghế xoay tựa đầu Kapvulix", "62x50x117-127 cm", 1_460_000),
    ("GX168/D1-168AB", "Ghế xoay tựa đầu Swaiper", "62x50x117-127 cm", 2_270_000),
    ("GXC99", "Ghế xoay ngả lưng, tựa đầu", "60x61x106-116 cm", 680_000),
    ("GDW090", "Ghế da quản lý, giám đốc", "57x48x115-123 cm", 870_000),
    ("GDW070", "Ghế giám đốc tay bạc", "50x67x108-115 cm", 1_020_000),
    ("GD9925", "Ghế giám đốc tay đen", "50x67x108-115 cm", 1_070_000),
    ("GDW88", "Ghế giám đốc tay gỗ", "50x67x108-115 cm", 1_270_000),
    ("GDD006D-1", "Ghế giám đốc có để chân", "66x52x110-117 cm", 1_900_000),
    ("GDW026", "Ghế giám đốc GDW026", "66x52x110-117 cm", 1_710_000),
    ("GDW026B", "Ghế giám đốc GDW026B", "66x52x110-117 cm", 2_050_000),
    ("CQ4001", "Ghế chân quỳ lưới CQ4001", "51x49x94 cm", 250_000),
    ("CQ4006", "Ghế chân quỳ CQ4006", "60x47x88-98 cm", 360_000),
    ("CQ4019", "Ghế chân quỳ lưới lưng thấp CQ4019", "56x49x93 cm", 290_000),
    ("CQ4020", "Ghế chân quỳ lưới lưng cao CQ4020", "56x49x105 cm", 300_000),
    ("CQ4009", "Ghế chân quỳ lưới lưng thấp CQ4009", "56x49x93 cm", 330_000),
    ("CQ4010", "Ghế chân quỳ lưới lưng cao CQ4010", "56x49x105 cm", 360_000),
    ("CQ6009", "Ghế chân quỳ lưới lưng thấp CQ6009", "56x49x93 cm", 380_000),
    ("CQ6010", "Ghế chân quỳ lưới lưng cao CQ6010", "44x47x105 cm", 420_000),
    ("CQ2001", "Ghế chân quỳ lưới lưng thấp CQ2001", "58x50x88 cm", 500_000),
    ("CQ2002", "Ghế chân quỳ lưới lưng cao CQ2002", "58x52x108 cm", 530_000),
    ("CQ4009B", "Ghế chân quỳ da lưng thấp CQ4009B", "58x52x91 cm", 420_000),
    ("CQ4010B", "Ghế chân quỳ da lưng cao CQ4010B", "58x52x108 cm", 430_000),
    ("CQ2003", "Ghế chân quỳ da lưng thấp CQ2003", "58x52x88 cm", 530_000),
    ("CQ2004", "Ghế chân quỳ lưới lưng cao CQ2004", "58x52x108 cm", 560_000),
    ("CQW118", "Ghế chân quỳ da CQW118", "58x69x99 cm", 710_000),
    ("Sofa1+1+3", "Bộ sofa da khung thép", "Bộ 1+1+3", 3_000_000),
]


def safe_name(code: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", code).strip("_")


def trim_and_frame(image: Image.Image) -> Image.Image:
    image = image.convert("RGB")
    background = Image.new("RGB", image.size, "white")
    diff = ImageChops.difference(image, background).convert("L")
    diff = diff.point(lambda value: 255 if value > 18 else 0)
    bbox = diff.getbbox()
    if bbox:
        image = image.crop(bbox)
    scale = min(1080 / image.width, 780 / image.height)
    image = image.resize(
        (max(1, round(image.width * scale)), max(1, round(image.height * scale))),
        Image.Resampling.LANCZOS,
    )
    canvas = Image.new("RGB", (1200, 900), "white")
    canvas.paste(image, ((1200 - image.width) // 2, (900 - image.height) // 2))
    return canvas


def save_pdf_image(document: fitz.Document, xref: int, destination: Path) -> None:
    payload = document.extract_image(xref)
    image = Image.open(BytesIO(payload["image"]))
    destination.parent.mkdir(parents=True, exist_ok=True)
    trim_and_frame(image).save(destination, "JPEG", quality=90, optimize=True)


def gevi_images(pdf_path: Path) -> dict[str, str]:
    document = fitz.open(pdf_path)
    selected = []
    for page_number, page in enumerate(document):
        candidates = []
        for info in page.get_image_info(xrefs=True):
            x0, y0, x1, y1 = info["bbox"]
            area = (x1 - x0) * (y1 - y0)
            if page_number == 0 and area > 5_000 and y0 > 80:
                candidates.append(info)
            elif page_number == 1 and y0 < 680 and area > 2_500:
                candidates.append(info)
            elif page_number == 2 and area > 2_500:
                candidates.append(info)
        candidates.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))
        selected.extend(candidates)
    if len(selected) != len(GEVI_PRODUCTS):
        raise RuntimeError(f"Expected {len(GEVI_PRODUCTS)} Gevi images, found {len(selected)}")

    result = {}
    for product, info in zip(GEVI_PRODUCTS, selected):
        code = product[0]
        relative = Path("assets") / "product_images" / "gevi" / f"{safe_name(code)}.jpg"
        save_pdf_image(document, info["xref"], PROJECT_ROOT / relative)
        result[code] = relative.as_posix()
    document.close()
    return result


def t8_images(pdf_path: Path) -> dict[str, str]:
    document = fitz.open(pdf_path)
    result = {}
    for code, xrefs in T8_IMAGE_XREFS.items():
        images = [Image.open(BytesIO(document.extract_image(xref)["image"])).convert("RGB") for xref in xrefs]
        if len(images) == 1:
            combined = images[0]
        else:
            for image in images:
                image.thumbnail((520, 700), Image.Resampling.LANCZOS)
            combined = Image.new("RGB", (sum(image.width for image in images) + 30, max(image.height for image in images)), "white")
            left = 0
            for image in images:
                combined.paste(image, (left, (combined.height - image.height) // 2))
                left += image.width + 30
        relative = Path("assets") / "product_images" / "eece_t8" / f"{safe_name(code)}.jpg"
        destination = PROJECT_ROOT / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        trim_and_frame(combined).save(destination, "JPEG", quality=90, optimize=True)
        result[code] = relative.as_posix()
    document.close()
    return result


def classify(name: str) -> tuple[str, str, str, str]:
    lowered = name.lower()
    if "ghế" in lowered:
        material = "Da công nghiệp" if "da" in lowered or "giám đốc" in lowered else "Lưới, vải và khung kim loại/nhựa"
        variant = "DA" if "da" in lowered or "giám đốc" in lowered else "LUOI"
        return "GHẾ", "OFFICE_CHAIR", material, variant
    if "sofa" in lowered:
        return "SOFA", "SOFA_SET", "Da công nghiệp và khung kim loại/gỗ", "DA"
    if "tủ" in lowered or "hộc" in lowered:
        return "TỦ", "OFFICE_STORAGE", "Gỗ công nghiệp phủ melamine", "MELAMINE"
    return "BÀN", "OFFICE_DESK", "Gỗ công nghiệp phủ melamine, khung kim loại", "MELAMINE"


def make_rows(products, images, source_group: str, source_url: str, image_source: str):
    rows = []
    for product in products:
        if len(product) == 4:
            code, name, size, price = product
            color = ""
        else:
            code, color, price = product
            name, size = f"Ghế văn phòng EECE {code}", ""
        category, subcategory, material, variant = classify(name)
        if color:
            variant = color
        description = f"{name}. Kích thước: {size}." if size else f"{name}. Màu: {color}. Liên hệ để xác nhận kích thước và tùy chọn hoàn thiện."
        search_url = f"https://www.google.com/search?q={quote_plus(code + ' ' + name)}"
        rows.append({
            "Code": code,
            "ProductName": name,
            "Size": size,
            "Category": category,
            "Variant": variant,
            "BasePrice": price,
            "SaleFactor": 1.2,
            "SalePrice": round(price * 1.2),
            "Status": "ACTIVE",
            "Search_URL": search_url,
            "Search_Keyword": f"{code} {name}",
            "Material": material,
            "SubCategory": subcategory,
            "Description": description,
            "CatalogPrice": price,
            "Price_Mode": "CATALOG",
            "Hotline": HOTLINE,
            "Image_Status": "FOUND",
            "Image_URL": images[code],
            "Image_Source": image_source,
            "Harvest_Status": "PDF_IMPORTED",
            "Source_Group": source_group,
            "Source_URL": source_url,
        })
    return rows


def write_source(path: Path, rows: list[dict]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "MASTER_V2"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row.get(header, "") for header in HEADERS])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(HEADERS, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 3, 14), 45)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def main() -> None:
    gevi_pdf = PDF_DIR / "BANG GIA GHE - Gevi Furniture.pdf"
    t8_pdf = PROJECT_ROOT / "exports" / "pdf_intake" / "t8" / "source_t8.pdf"
    for path in (gevi_pdf, t8_pdf):
        if not path.is_file():
            raise FileNotFoundError(path)

    gevi_rows = make_rows(
        GEVI_PRODUCTS, gevi_images(gevi_pdf), "GEVI_GHE", "https://gevi.vn",
        "PDF: BANG GIA GHE - Gevi Furniture.pdf",
    )
    t8_rows = make_rows(
        T8_PRODUCTS, t8_images(t8_pdf), "EECE_T8",
        "", "PDF: Bảng giá chuẩn T8.pdf",
    )
    write_source(PROJECT_ROOT / "GEVI_GHE" / "GEVI_GHE_FINAL_V2.xlsx", gevi_rows)
    write_source(PROJECT_ROOT / "EECE_T8" / "EECE_T8_FINAL_V2.xlsx", t8_rows)
    print(f"Created {len(gevi_rows) + len(t8_rows)} source rows and product images.")


if __name__ == "__main__":
    main()
