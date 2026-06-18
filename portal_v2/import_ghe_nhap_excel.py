from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from urllib.parse import quote_plus

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = PROJECT_ROOT / "exports" / "incoming" / "ghe_nhap_phu_vuong.xlsx"
IMAGE_DIR = PROJECT_ROOT / "assets" / "product_images" / "ghe_nhap"
OUTPUT_WORKBOOK = PROJECT_ROOT / "GHE_NHAP" / "GHE_NHAP_FINAL_V2.xlsx"
SOURCE_GROUP = "GHE_NHAP"
HOTLINE = "0929878666"
HEADERS = [
    "Code", "ProductName", "Size", "Category", "Variant", "BasePrice",
    "SaleFactor", "SalePrice", "Status", "Search_URL", "Search_Keyword",
    "Material", "SubCategory", "Description", "CatalogPrice", "Price_Mode",
    "Hotline", "Image_Status", "Image_URL", "Image_Source", "Harvest_Status",
    "Source_Group", "Source_URL",
]


def clean_text(value: object) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def safe_code(value: object) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"\s+", "-", text)
    text = text.replace("NAN", "NAN")
    return text.strip("-")


def is_probable_color(value: object) -> bool:
    text = clean_text(value).casefold()
    if not text:
        return True
    color_words = [
        "đen", "trắng", "nâu", "xám", "ghi", "cam", "vàng", "xanh", "đỏ",
        "kem", "bạc", "gỗ", "mint", "không đệm", "ko đệm",
    ]
    return any(word in text for word in color_words)


def choose_code_and_color(value_a: object, value_b: object) -> tuple[str, str]:
    a, b = clean_text(value_a), clean_text(value_b)
    if is_probable_color(a) and not is_probable_color(b):
        return safe_code(b), a
    if is_probable_color(b) and not is_probable_color(a):
        return safe_code(a), b
    if re.search(r"\d", b or ""):
        return safe_code(b), a
    return safe_code(a or b), b if a else ""


def image_filename(code: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", code).strip("_") + ".jpg"


def trim_and_frame(image: Image.Image) -> Image.Image:
    image = image.convert("RGB")
    background = Image.new("RGB", image.size, "white")
    diff = ImageChops.difference(image, background).convert("L")
    diff = diff.point(lambda value: 255 if value > 18 else 0)
    bbox = diff.getbbox()
    if bbox:
        image = image.crop(bbox)
    scale = min(1080 / image.width, 780 / image.height)
    image = image.resize(
        (max(1, round(image.width * scale)), max(1, round(image.height * scale))),
        Image.Resampling.LANCZOS,
    )
    canvas = Image.new("RGB", (1200, 900), "white")
    canvas.paste(image, ((1200 - image.width) // 2, (900 - image.height) // 2))
    return canvas


def save_row_image(code: str, images: list[Image.Image]) -> str:
    if not images:
        return ""
    prepared = []
    for image in images:
        image = image.convert("RGB")
        image.thumbnail((520, 760), Image.Resampling.LANCZOS)
        prepared.append(image)
    if len(prepared) == 1:
        combined = prepared[0]
    else:
        combined = Image.new(
            "RGB",
            (sum(image.width for image in prepared) + 30 * (len(prepared) - 1), max(image.height for image in prepared)),
            "white",
        )
        left = 0
        for image in prepared:
            combined.paste(image, (left, (combined.height - image.height) // 2))
            left += image.width + 30
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    relative = Path("assets") / "product_images" / "ghe_nhap" / image_filename(code)
    trim_and_frame(combined).save(PROJECT_ROOT / relative, "JPEG", quality=90, optimize=True)
    return relative.as_posix()


def fallback_image(code: str) -> str:
    for folder in ("eece_t8", "gevi"):
        for candidate in (IMAGE_DIR.parent / folder).glob(f"*{image_filename(code)}"):
            if candidate.is_file():
                IMAGE_DIR.mkdir(parents=True, exist_ok=True)
                relative = Path("assets") / "product_images" / "ghe_nhap" / image_filename(code)
                Image.open(candidate).convert("RGB").save(PROJECT_ROOT / relative, "JPEG", quality=90, optimize=True)
                return relative.as_posix()
    return ""


def extract_images_by_row(worksheet) -> dict[int, list[Image.Image]]:
    images_by_row: dict[int, list[Image.Image]] = {}
    for embedded in worksheet._images:
        row = embedded.anchor._from.row + 1
        width = int(getattr(embedded, "width", 0) or 0)
        height = int(getattr(embedded, "height", 0) or 0)
        if width < 8 or height < 8:
            continue
        raw = embedded._data()
        images_by_row.setdefault(row, []).append(Image.open(BytesIO(raw)).convert("RGB"))
    return images_by_row


def classify(name: str, material: str) -> tuple[str, str, str]:
    text = f"{name} {material}".casefold()
    if "da" in text:
        variant = "DA"
    elif "lưới" in text:
        variant = "LUOI"
    else:
        variant = "GHE"
    return "GHẾ", "OFFICE_CHAIR", variant


def numeric(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^\d.]", "", str(value))
    return float(text) if text else 0.0


def build_rows() -> list[dict[str, object]]:
    if not SOURCE_WORKBOOK.is_file():
        raise FileNotFoundError(SOURCE_WORKBOOK)
    workbook = load_workbook(SOURCE_WORKBOOK, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    images_by_row = extract_images_by_row(worksheet)

    imported: list[dict[str, object]] = []
    prior_image_by_code: dict[str, str] = {}
    for row_index in range(3, worksheet.max_row + 1):
        name = clean_text(worksheet.cell(row_index, 1).value)
        if not name:
            continue
        code, color = choose_code_and_color(worksheet.cell(row_index, 3).value, worksheet.cell(row_index, 4).value)
        if not code:
            continue
        price = numeric(worksheet.cell(row_index, 5).value)
        if price <= 0:
            continue
        size = clean_text(worksheet.cell(row_index, 8).value)
        material = clean_text(worksheet.cell(row_index, 9).value)
        description = clean_text(worksheet.cell(row_index, 10).value)
        category, subcategory, inferred_variant = classify(name, material)
        image_url = save_row_image(code, images_by_row.get(row_index, []))
        if not image_url and code in prior_image_by_code:
            image_url = prior_image_by_code[code]
        if not image_url:
            image_url = fallback_image(code)
        if image_url:
            prior_image_by_code[code] = image_url
        search = f"https://www.google.com/search?q={quote_plus(code + ' ' + name)}"
        imported.append({
            "Code": code,
            "ProductName": name,
            "Size": size,
            "Category": category,
            "Variant": color or inferred_variant,
            "BasePrice": round(price),
            "SaleFactor": 1.2,
            "SalePrice": round(price * 1.2),
            "Status": "ACTIVE",
            "Search_URL": search,
            "Search_Keyword": f"{code} {name}",
            "Material": material,
            "SubCategory": subcategory,
            "Description": description if description else f"{name}. Màu: {color}.",
            "CatalogPrice": round(price),
            "Price_Mode": "CATALOG",
            "Hotline": HOTLINE,
            "Image_Status": "FOUND" if image_url else "MISSING",
            "Image_URL": image_url,
            "Image_Source": "Excel: Ghế Nhập Phú Vương.xlsx",
            "Harvest_Status": "EXCEL_IMPORTED",
            "Source_Group": SOURCE_GROUP,
            "Source_URL": search,
        })

    selected: dict[str, dict[str, object]] = {}
    order: list[str] = []
    for row in imported:
        key = str(row["Code"]).casefold()
        if key not in selected:
            selected[key] = row
            order.append(key)
            continue
        current_price = numeric(selected[key]["BasePrice"])
        candidate_price = numeric(row["BasePrice"])
        if candidate_price > current_price:
            selected[key] = row
    return [selected[key] for key in order]


def write_workbook(rows: list[dict[str, object]]) -> None:
    OUTPUT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "MASTER_V2"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row.get(header, "") for header in HEADERS])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(HEADERS, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 3, 14), 45)
    workbook.save(OUTPUT_WORKBOOK)
    workbook.close()


def main() -> None:
    rows = build_rows()
    write_workbook(rows)
    missing = sum(1 for row in rows if not str(row.get("Image_URL", "")).strip())
    print(f"Created {len(rows)} GHE_NHAP rows from Excel. Missing images: {missing}.")


if __name__ == "__main__":
    main()
