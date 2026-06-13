from __future__ import annotations

from io import BytesIO
from pathlib import Path
from textwrap import wrap
from urllib.parse import urlparse

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont, ImageOps


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FONT_REGULAR_CANDIDATES = [
    Path(r"C:\Windows\Fonts\arial.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
]
FONT_BOLD_CANDIDATES = [
    Path(r"C:\Windows\Fonts\arialbd.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
]
HOTLINE = "0929.878.666"


def _font(size: int, bold: bool = False):
    candidates = FONT_BOLD_CANDIDATES if bold else FONT_REGULAR_CANDIDATES
    path = next((candidate for candidate in candidates if candidate.exists()), None)
    return ImageFont.truetype(str(path), size) if path else ImageFont.load_default()


def _image_bytes(source: object) -> bytes | None:
    value = str(source or "").strip()
    if not value:
        return None
    parsed = urlparse(value)
    try:
        if parsed.scheme in {"http", "https"}:
            response = requests.get(value, timeout=12, headers={"User-Agent": "ProductDB-V2/1.0"})
            response.raise_for_status()
            return response.content
        candidates = [Path(value), PROJECT_ROOT / value, PROJECT_ROOT.parent / value]
        path = next((candidate for candidate in candidates if candidate.is_file()), None)
        return path.read_bytes() if path else None
    except (OSError, requests.RequestException):
        return None


def _thumbnail(source: object, size: tuple[int, int]) -> Image.Image:
    raw = _image_bytes(source)
    if raw:
        try:
            image = Image.open(BytesIO(raw)).convert("RGB")
            return ImageOps.contain(image, size, Image.Resampling.LANCZOS)
        except (OSError, ValueError):
            pass
    image = Image.new("RGB", size, "#f2f4f7")
    draw = ImageDraw.Draw(image)
    draw.text((12, size[1] // 2 - 10), "Chưa có ảnh", fill="#667085", font=_font(16))
    return image


def _money(value: object) -> str:
    try:
        return f"{float(value):,.0f} đ".replace(",", ".")
    except (TypeError, ValueError):
        return "0 đ"


def export_excel(customer: dict[str, str], quotation: pd.DataFrame) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bao_gia"
    sheet.sheet_view.showGridLines = False
    info = [
        ("BÁO GIÁ PRODUCTDB V2", ""),
        ("Khách hàng", customer["Name"]),
        ("Điện thoại", customer["Phone"]),
        ("Kênh", customer["Channel"]),
        ("Ghi chú đơn hàng", customer["Notes"]),
        ("Hotline", HOTLINE),
    ]
    for row, (label, value) in enumerate(info, start=1):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="B42318")
    sheet.merge_cells("A1:K1")

    headers = ["Ảnh", "Mã", "Sản phẩm", "Kích thước", "Vật liệu", "Giá dữ liệu",
               "Đơn giá sale", "SL", "CK %", "Thành tiền", "Ghi chú"]
    header_row = 8
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="344054")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_buffers = []
    for offset, (_, item) in enumerate(quotation.iterrows(), start=1):
        row = header_row + offset
        values = [
            "", item.get("Code", ""), item.get("ProductName", ""), item.get("Size", ""),
            item.get("Material", ""), item.get("BasePrice", 0), item.get("UnitPrice", 0),
            item.get("Quantity", 1), item.get("Discount", 0), item.get("LineTotal", 0),
            item.get("ItemNote", ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in (6, 7, 10):
            sheet.cell(row, column).number_format = '#,##0 "đ"'
        thumbnail = _thumbnail(item.get("Image_URL", ""), (120, 80))
        buffer = BytesIO()
        thumbnail.save(buffer, format="PNG")
        buffer.seek(0)
        image_buffers.append(buffer)
        excel_image = ExcelImage(buffer)
        excel_image.width, excel_image.height = thumbnail.size
        sheet.add_image(excel_image, f"A{row}")
        sheet.row_dimensions[row].height = 64

    widths = [19, 16, 38, 22, 22, 16, 16, 8, 10, 18, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A9"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_png(customer: dict[str, str], quotation: pd.DataFrame) -> bytes:
    width, row_height = 1240, 190
    height = 380 + max(len(quotation), 1) * row_height + 170
    canvas = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(canvas)
    draw.rectangle((0, 0, width, 115), fill="#B42318")
    draw.text((55, 30), "BÁO GIÁ PRODUCTDB V2", fill="white", font=_font(38, True))
    draw.text((55, 135), f"Khách hàng: {customer['Name']}", fill="#101828", font=_font(23, True))
    draw.text((55, 172), f"Điện thoại: {customer['Phone']}   |   Kênh: {customer['Channel']}", fill="#344054", font=_font(20))
    draw.text((55, 208), f"Ghi chú: {customer['Notes'] or '-'}", fill="#344054", font=_font(19))
    draw.line((55, 255, width - 55, 255), fill="#D0D5DD", width=2)

    y = 280
    total = 0.0
    for _, item in quotation.iterrows():
        thumbnail = _thumbnail(item.get("Image_URL", ""), (170, 140))
        canvas.paste(thumbnail, (55, y + 10))
        x = 250
        draw.text((x, y + 5), str(item.get("Code", "")), fill="#B42318", font=_font(20, True))
        name_lines = wrap(str(item.get("ProductName", "")), width=48)[:2]
        draw.multiline_text((x, y + 35), "\n".join(name_lines), fill="#101828", font=_font(22, True), spacing=5)
        detail_y = y + 95
        draw.text((x, detail_y), f"SL: {int(item.get('Quantity', 1))}   CK: {float(item.get('Discount', 0)):g}%", fill="#475467", font=_font(18))
        draw.text((760, y + 25), f"Đơn giá: {_money(item.get('UnitPrice', 0))}", fill="#101828", font=_font(20))
        draw.text((760, y + 65), f"Thành tiền: {_money(item.get('LineTotal', 0))}", fill="#B42318", font=_font(22, True))
        note = str(item.get("ItemNote", "")).strip()
        if note:
            draw.text((x, y + 130), "Ghi chú: " + note[:75], fill="#667085", font=_font(17))
        draw.line((55, y + row_height - 8, width - 55, y + row_height - 8), fill="#EAECF0", width=2)
        total += float(item.get("LineTotal", 0) or 0)
        y += row_height

    draw.text((700, y + 20), "TỔNG CỘNG", fill="#101828", font=_font(26, True))
    draw.text((width - 55, y + 20), _money(total), fill="#B42318", font=_font(28, True), anchor="ra")
    draw.text((55, y + 90), "Sản phẩm có thể thay đổi kích thước, vật liệu và màu sắc theo yêu cầu.", fill="#344054", font=_font(19))
    draw.text((width - 55, y + 90), f"Hotline: {HOTLINE}", fill="#B42318", font=_font(22, True), anchor="ra")
    output = BytesIO()
    canvas.save(output, format="PNG", optimize=True)
    return output.getvalue()


def export_pdf(customer: dict[str, str], quotation: pd.DataFrame) -> bytes:
    png = export_png(customer, quotation)
    image = Image.open(BytesIO(png)).convert("RGB")
    output = BytesIO()
    image.save(output, format="PDF", resolution=150.0)
    return output.getvalue()
